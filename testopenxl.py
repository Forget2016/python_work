#/usr/bash/python
# *--coding:uft-8 --*

"""the code is just testing the package of openpyxl"""

from openpyxl import Workbook

#creating the excel file,if want to use the format inference
#need to use the parameter guess_types
wb = Workbook()
wb = Workbook(guess_types=True) #allow use the type and format

#the other way is loading the file
wb = load_workbook("testfile.xlsx")

#the second step create the sheet,insert at the end
ws = wb.create_sheet("lastsheet")
#the other way insert the sheet at the first position
ws = wb.create_sheet("firstsheet", 0)

#change the title by using the title property of sheet
ws.title = "changetitle"


#the cell of sheet actives, for example change the A4 value
ws("A4") = 12
ws.cell(row=4, column=1, value=12)